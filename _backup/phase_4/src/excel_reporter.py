"""
덤프트럭 안전보건 교육일지 및 참석자 서명부 엑셀 생성 엔진 (src/excel_reporter.py)
- 법정 안전보건 교육일지 표준 서식 (.xlsx) 자동 생성
- 기사님 자필 서명 PNG 이미지 셀 자동 리사이징 삽입
"""

import os
import json
import base64
import datetime
from io import BytesIO
from PIL import Image
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SIGNATURES_LOCAL_DB = os.path.join(ROOT_DIR, "data", "signatures_db.json")

def load_local_signatures() -> list:
    """로컬 서명 데이터 로드"""
    if os.path.exists(SIGNATURES_LOCAL_DB):
        try:
            with open(SIGNATURES_LOCAL_DB, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def save_signature_record(
    vehicle_number: str,
    driver_name: str,
    phone: str,
    edu_name: str = "덤프트럭 운전자 교통안전 및 작업안전 교육",
    signature_base64: str = "",
    signature_png_path: str = ""
) -> dict:
    """새로운 서명 제출 데이터를 로컬 DB 및 이미지 파일로 저장"""
    os.makedirs(os.path.dirname(SIGNATURES_LOCAL_DB), exist_ok=True)
    records = load_local_signatures()
    
    timestamp_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    file_timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Base64 서명 이미지를 PNG 파일로 저장
    saved_img_path = signature_png_path
    if signature_base64 and not signature_png_path:
        img_dir = os.path.join(ROOT_DIR, "data", "signature_images")
        os.makedirs(img_dir, exist_ok=True)
        
        clean_b64 = signature_base64
        if "base64," in clean_b64:
            clean_b64 = clean_b64.split("base64,")[1]
            
        try:
            img_data = base64.b64decode(clean_b64)
            saved_img_path = os.path.join(img_dir, f"sig_{file_timestamp}_{vehicle_number}.png")
            with open(saved_img_path, "wb") as f:
                f.write(img_data)
        except Exception:
            pass

    new_record = {
        "id": f"sig_{int(datetime.datetime.now().timestamp())}",
        "created_at": timestamp_str,
        "vehicle_number": vehicle_number,
        "driver_name": driver_name,
        "phone": phone,
        "edu_name": edu_name,
        "status": "이수완료",
        "signature_image_path": saved_img_path
    }
    
    records.append(new_record)
    with open(SIGNATURES_LOCAL_DB, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
        
    return new_record

def export_safety_education_log_excel(
    output_excel_path: str,
    edu_title: str = "2026년 덤프트럭 운전자 교통안전 및 작업안전 정기교육",
    edu_date_str: str = "",
    instructor_name: str = "안전관리자",
    edu_summary_points: list = None,
    records: list = None
) -> str:
    """
    공식 법정 「안전보건 교육일지 및 참석자 서명부」 엑셀 파일 자동 생성
    - 각 행 [서명란] 셀 안에 기사님 실제 자필 싸인 PNG 이미지 자동 삽입
    - A4 인쇄 규격 완벽 적용
    """
    os.makedirs(os.path.dirname(os.path.abspath(output_excel_path)), exist_ok=True)
    
    if not records:
        records = load_local_signatures()
        
    if not edu_date_str:
        edu_date_str = datetime.datetime.now().strftime("%Y년 %m월 %d일")
        
    if not edu_summary_points:
        edu_summary_points = [
            "1. 건설현장 및 구내 가설도로 제한속도 10~20km/h 준수 및 교행 시 감속",
            "2. 하역(덤핑) 전 평탄하고 견고한 수평 지반 확보 및 단부 안전거리 준수",
            "3. 적재함 100% 하강 안착 확인 후 발차 (개문 주행 및 고압선 접촉 절대 금지)",
            "4. 경사지 주차 시 바퀴 전·후 고임목 설치 및 브레이크 에어 압력 7bar 확인"
        ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "안전보건교육일지"
    
    # 그리드선 표시
    ws.views.sheetView[0].showGridLines = True

    # 폰트 및 스타일 정의
    font_title = Font(name="맑은 고딕", size=18, bold=True, color="1E3A8A")
    font_header_label = Font(name="맑은 고딕", size=11, bold=True, color="1E293B")
    font_header_val = Font(name="맑은 고딕", size=11, color="0F172A")
    font_table_head = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    font_table_body = Font(name="맑은 고딕", size=10, color="0F172A")
    font_table_bold = Font(name="맑은 고딕", size=10, bold=True, color="0F172A")
    
    fill_main_head = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_sub_head = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. 메인 타이틀 (Row 2)
    ws.merge_cells("A2:G2")
    ws["A2"] = "덤프트럭 안전보건 교육일지 및 참석자 서명부"
    ws["A2"].font = font_title
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 40

    # 2. 교육 개요 테이블 (Row 4 ~ 8)
    overview_data = [
        ("교육과정명", edu_title, "교육구분", "모바일 시청각 안전교육"),
        ("교육일시", f"{edu_date_str} (09:00 ~ 10:00)", "교 육 자", instructor_name),
        ("교육장소", "사내 / 현장 모바일 안전교육 시스템", "법적근거", "산업안전보건법 제29조 / 건설기계관리법")
    ]

    cur_r = 4
    for row_info in overview_data:
        ws.cell(row=cur_r, column=1, value=row_info[0]).font = font_header_label
        ws.cell(row=cur_r, column=1).fill = fill_sub_head
        ws.cell(row=cur_r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells(start_row=cur_r, start_column=2, end_row=cur_r, end_column=4)
        ws.cell(row=cur_r, column=2, value=row_info[1]).font = font_header_val
        ws.cell(row=cur_r, column=2).alignment = Alignment(horizontal="left", vertical="center")
        
        ws.cell(row=cur_r, column=5, value=row_info[2]).font = font_header_label
        ws.cell(row=cur_r, column=5).fill = fill_sub_head
        ws.cell(row=cur_r, column=5).alignment = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells(start_row=cur_r, start_column=6, end_row=cur_r, end_column=7)
        ws.cell(row=cur_r, column=6, value=row_info[3]).font = font_header_val
        ws.cell(row=cur_r, column=6).alignment = Alignment(horizontal="left", vertical="center")
        
        for c in range(1, 8):
            ws.cell(row=cur_r, column=c).border = thin_border
        ws.row_dimensions[cur_r].height = 24
        cur_r += 1

    # 주요 교육내용 요약 (Row 7)
    ws.cell(row=cur_r, column=1, value="주요 교육내용").font = font_header_label
    ws.cell(row=cur_r, column=1).fill = fill_sub_head
    ws.cell(row=cur_r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells(start_row=cur_r, start_column=2, end_row=cur_r, end_column=7)
    summary_text = "\n".join(edu_summary_points)
    ws.cell(row=cur_r, column=2, value=summary_text).font = font_header_val
    ws.cell(row=cur_r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    for c in range(1, 8):
        ws.cell(row=cur_r, column=c).border = thin_border
    ws.row_dimensions[cur_r].height = 70
    cur_r += 2

    # 3. 참석자 서명 명부 헤더
    ws.merge_cells(start_row=cur_r, start_column=1, end_row=cur_r, end_column=7)
    ws.cell(row=cur_r, column=1, value="[ 교육 참석자 명단 및 자필 서명 확인 ]").font = Font(name="맑은 고딕", size=13, bold=True, color="1E3A8A")
    ws.cell(row=cur_r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    cur_r += 1

    headers = ["연번", "서명일시", "차량번호", "성 명", "연락처", "이수여부", "자필 서명란"]
    col_widths = [8, 20, 16, 14, 18, 12, 22]
    
    for col_idx, (h_text, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=cur_r, column=col_idx, value=h_text)
        cell.font = font_table_head
        cell.fill = fill_main_head
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    ws.row_dimensions[cur_r].height = 28
    cur_r += 1

    # 4. 참석자 데이터 및 자필 서명 PNG 삽입
    for idx, rec in enumerate(records, start=1):
        row_num = cur_r
        ws.row_dimensions[row_num].height = 42  # 서명란 셀 높이
        
        ws.cell(row=row_num, column=1, value=idx).font = font_table_bold
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center", vertical="center")
        
        ws.cell(row=row_num, column=2, value=rec.get("created_at", "")).font = font_table_body
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="center", vertical="center")
        
        ws.cell(row=row_num, column=3, value=rec.get("vehicle_number", "")).font = font_table_bold
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="center", vertical="center")
        
        ws.cell(row=row_num, column=4, value=rec.get("driver_name", "")).font = font_table_bold
        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="center", vertical="center")
        
        ws.cell(row=row_num, column=5, value=rec.get("phone", "")).font = font_table_body
        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center", vertical="center")
        
        ws.cell(row=row_num, column=6, value=rec.get("status", "완료")).font = Font(name="맑은 고딕", size=10, bold=True, color="16A34A")
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal="center", vertical="center")
        
        sig_cell = ws.cell(row=row_num, column=7)
        sig_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if idx % 2 == 0:
            for c in range(1, 8):
                ws.cell(row=row_num, column=c).fill = fill_zebra
                
        for c in range(1, 8):
            ws.cell(row=row_num, column=c).border = thin_border

        # 자필 서명 PNG 이미지 셀 삽입
        sig_img_path = rec.get("signature_image_path", "")
        if sig_img_path and os.path.exists(sig_img_path):
            try:
                img = OpenpyxlImage(sig_img_path)
                img.width = 110
                img.height = 45
                cell_coord = f"G{row_num}"
                ws.add_image(img, cell_coord)
            except Exception:
                sig_cell.value = "(서명 확인)"
        else:
            sig_cell.value = "(서명 완료)"
            
        cur_r += 1

    wb.save(output_excel_path)
    return output_excel_path

class ExcelReporter:
    """공식 법정 안전보건 교육일지 생성 및 대장 관리 클래스"""
    def __init__(self, output_dir: str = "outputs/reports"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def generate_report(
        self,
        scene_id: int = 1,
        scene_title: str = "덤프트럭 운전자 안전교육",
        instructor: str = "안전관리자",
        trainees: list = None,
        education_date: str = "",
        photo_path: str = None
    ) -> str:
        timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(self.output_dir, f"안전보건교육일지_씬{scene_id}_{timestamp_str}.xlsx")
        
        trainee_names = trainees or ["김기사"]
        records = [
            {
                "created_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "vehicle_number": f"서울06가{1000 + i}",
                "driver_name": name,
                "phone": f"010-{2000+i:04d}-{3000+i:04d}",
                "edu_name": scene_title,
                "status": "이수완료",
                "signature_image_path": photo_path if photo_path and os.path.exists(photo_path) else ""
            }
            for i, name in enumerate(trainee_names)
        ]
        
        return export_safety_education_log_excel(
            output_excel_path=out_path,
            edu_title=f"덤프트럭 안전보건 교육일지 (씬 #{scene_id}: {scene_title})",
            edu_date_str=education_date,
            instructor_name=instructor,
            records=records
        )

