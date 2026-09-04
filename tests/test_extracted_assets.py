"""
추출된 핵심 자산 무결성 단위 테스트 (tests/test_extracted_assets.py)
"""

import os
import unittest
import openpyxl
from src.prompt_builder import PromptBuilder, INVARIANT_CORE_TRUCK, SITE_PRESETS
from src.audio_master import find_korean_font, VOICES, DEFAULT_VOICE
from src.excel_reporter import export_safety_education_log_excel

class TestExtractedAssets(unittest.TestCase):
    def test_prompt_builder_dynamic(self):
        builder = PromptBuilder()
        scene = {
            "site_preset": "embankment_dump",
            "cargo_state": "dumping_lift",
            "weather_preset": "heavy_rain_mud",
            "role_preset": "spotter_guide",
            "scene_block": "사토장 연약지반 덤핑 중 전도 방지"
        }
        prompt_res = builder.build_2d_prompt(scene)
        self.assertIn("Korean cab-over", prompt_res["positive"])
        self.assertIn("reclamation embankment", prompt_res["positive"])
        self.assertIn("hydraulic cylinder", prompt_res["positive"])
        self.assertIn("heavy pouring rain", prompt_res["positive"])
        self.assertIn("no speech bubble", prompt_res["positive"])

        h3_prompt = builder.build_h3_motion_prompt(scene)
        self.assertIn("<Picture 1>", h3_prompt)
        self.assertIn("at 0.00 seconds into the target video", h3_prompt)
        print("PromptBuilder test passed!")

    def test_audio_master_font_and_voices(self):
        font_p = find_korean_font()
        self.assertTrue(os.path.exists(font_p) or "malgun" in font_p)
        self.assertIn("ko-KR-InJoonNeural", VOICES.values())
        print("AudioMaster test passed!")

    def test_excel_reporter_generation(self):
        test_out = "outputs/reports/test_report.xlsx"
        os.makedirs("outputs/reports", exist_ok=True)
        res_path = export_safety_education_log_excel(
            test_out,
            edu_title="테스트 덤프트럭 안전교육",
            records=[{
                "created_at": "2026-09-04 10:00:00",
                "vehicle_number": "경기88바1234",
                "driver_name": "홍길동",
                "phone": "010-1234-5678",
                "status": "이수완료",
                "signature_image_path": ""
            }]
        )
        self.assertTrue(os.path.exists(res_path))
        wb = openpyxl.load_workbook(res_path)
        self.assertIn("안전보건교육일지", wb.sheetnames)
        ws = wb["안전보건교육일지"]
        self.assertEqual(ws["A2"].value, "덤프트럭 안전보건 교육일지 및 참석자 서명부")
        print("ExcelReporter test passed!")

    def test_h3_workflow_json_exists(self):
        wf_path = "workflows/Minimax H3 Ref2VA + Easy Cache _ Spectrum + Sage.json"
        self.assertTrue(os.path.exists(wf_path))
        print("H3 Ref2VA workflow verified!")

if __name__ == "__main__":
    unittest.main()
